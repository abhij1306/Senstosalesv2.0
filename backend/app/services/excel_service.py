"""
Excel Generation Service
Uses XlsxWriter to generate formatted Excel reports with strict layout control.
"""

import io
import logging
import sqlite3
from typing import Dict, List

import pandas as pd
import xlsxwriter
from fastapi.responses import StreamingResponse

from app.core.num_to_words import amount_to_words

logger = logging.getLogger(__name__)


class ExcelService:
    @staticmethod
    def generate_response(data: List[Dict], report_type: str) -> StreamingResponse:
        """
        Convert list of dicts to Excel download response (Legacy fallback)
        """
        output = io.BytesIO()

        # Convert to DataFrame
        if not data:
            df = pd.DataFrame()  # Empty
        else:
            df = pd.DataFrame(data)

        # Write to Excel with formatting
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Report", index=False)

            workbook = writer.book
            worksheet = writer.sheets["Report"]

            # Formats
            header_fmt = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#4F81BD",
                    "font_color": "white",
                    "border": 1,
                }
            )

            # Apply header format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_fmt)

            # Auto-adjust column width
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, column_len)

        output.seek(0)

        filename = f"{report_type}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @staticmethod
    def _write_standard_header(
        worksheet,
        workbook,
        columns: int,
        db: sqlite3.Connection,
        title: str = None,
        layout: str = "invoice",
        font_name: str = "Calibri",
    ):
        """
        Consistently writes the business header across all reports with layout options.
        layout: 'invoice' (Standard for Sales Invoice) or 'challan' (Standard for DC, Summary, GC)
        """
        # Fetch settings from DB
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception as e:
            logger.error(f"Failed to fetch business settings, using defaults: {e}")
            settings = {}

        # Default Fallbacks
        # Default Fallbacks - STRICTLY from Settings
        s_name = settings.get("supplier_name", "")
        s_desc = settings.get("supplier_description", "")
        s_addr = settings.get("supplier_address", "")
        s_gst = settings.get("supplier_gstin", "")
        s_phone = settings.get("supplier_contact", "")
        s_state = settings.get("supplier_state", "")
        s_state_code = settings.get("supplier_state_code", "")

        # Formats
        title_fmt = workbook.add_format(
            {"bold": True, "font_size": 18, "align": "center", "font_name": font_name}
        )
        subtitle_fmt = workbook.add_format(
            {"bold": True, "font_size": 10, "align": "center", "font_name": font_name}
        )
        tel_fmt = workbook.add_format(
            {"font_size": 10, "align": "center", "font_name": font_name}
        )
        name_fmt = workbook.add_format(
            {"bold": True, "font_size": 14, "align": "left", "font_name": font_name}
        )
        detail_fmt = workbook.add_format(
            {"font_size": 11, "align": "left", "font_name": font_name}
        )
        bold_detail = workbook.add_format(
            {"bold": True, "font_size": 11, "align": "left", "font_name": font_name}
        )

        row = 0

        if layout == "invoice":
            # Layout matching 'GST_INV_11.xls'
            if title:
                worksheet.merge_range(
                    row,
                    0,
                    row,
                    columns - 1,
                    title,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 14,
                            "align": "center",
                            "font_name": font_name,
                        }
                    ),
                )
                row += 2  # Add spacing

            worksheet.merge_range(row, 0, row, 7, s_name, name_fmt)  # Col H is index 7
            row += 1
            worksheet.merge_range(row, 0, row + 1, 7, s_addr, detail_fmt)
            row += 2
            worksheet.merge_range(row, 0, row, 7, f"GSTIN/UIN: {s_gst}", bold_detail)
            row += 1
            worksheet.merge_range(
                row,
                0,
                row,
                7,
                f"State Name : {s_state}, Code : {s_state_code}",
                bold_detail,
            )
            row += 1
            worksheet.merge_range(row, 0, row, 7, f"Contact : {s_phone}", bold_detail)
            row += 1

        elif layout == "challan":
            # Layout matching 'DC12.xls'
            # Row 1: Tel (Left), GSTIN (Right)
            worksheet.write(row, 0, f"Tel. No. {s_phone}", tel_fmt)
            worksheet.merge_range(
                row, columns - 1, row, columns - 1, f"GSTIN: {s_gst}", tel_fmt
            )
            row += 2

            # Branding (Rows 3, 4, 5)
            worksheet.merge_range(row, 0, row, columns - 1, s_name, title_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, columns - 1, s_desc, subtitle_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, columns - 1, s_addr, subtitle_fmt)
            row += 2

            if title:
                worksheet.merge_range(
                    row,
                    0,
                    row,
                    columns - 1,
                    title,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 14,
                            "align": "center",
                            "font_name": font_name,
                        }
                    ),
                )
                row += 1

        return row

    @staticmethod
    def _write_buyer_block(
        worksheet,
        workbook,
        row: int,
        col: int,
        db: sqlite3.Connection,
        header: Dict = None,
        width: int = 5,
        label: str = "Buyer :",
        font_name: str = "Calibri",
    ):
        """
        Consistently writes the Buyer/Consignee block.
        Fetches from DB settings as default, overriden by specific record header if available.
        """
        # Fetch Supplier Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            {row["key"]: row["value"] for row in rows}
        except Exception as e:
            logger.error(f"Failed to fetch business settings for buyer block: {e}")

        # Fetch Default Buyer if not provided in header
        default_buyer = {}
        if not header.get("consignee_name"):
            try:
                buyer_row = db.execute(
                    "SELECT name, billing_address, gstin, place_of_supply FROM buyers WHERE is_default = 1 AND is_active = 1 LIMIT 1"
                ).fetchone()
                if buyer_row:
                    default_buyer = dict(buyer_row)
            except Exception as e:
                logger.error(f"Failed to fetch default buyer: {e}")

        # Default Buyer Info logic: Header > Default Buyer > Settings > Empty
        b_name = header.get("consignee_name") or default_buyer.get("name") or settings.get("buyer_name", "")
        b_addr = header.get("consignee_address") or default_buyer.get("billing_address") or settings.get("buyer_address", "")
        b_gst = header.get("consignee_gstin") or default_buyer.get("gstin") or settings.get("buyer_gstin", "")
        
        # Parse Place of Supply if needed
        b_pos_raw = header.get("place_of_supply") or default_buyer.get("place_of_supply") or "BHOPAL, MP"
        b_pos = b_pos_raw
        
        # State logic (simple extraction if not provided)
        b_state = header.get("buyer_state") or "MP"
        if not header.get("buyer_state") and default_buyer.get("place_of_supply"):
             # Simple heuristic: last word or known states. For now defaulting to MP or extracting from POS
             pass

        # Formats - ALL buyer details should be BOLD with borders
        bold_border_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "font_name": font_name,
                "border": 1,
                "valign": "vcenter",
            }
        )

        if label:
            worksheet.merge_range(row, col, row, col + width, label, bold_border_fmt)
            row += 1

        # Each line should be in its own row with borders - NO merging
        worksheet.merge_range(row, col, row, col + width, b_name, bold_border_fmt)
        row += 1

        # Address might be multi-line but still one row with border
        worksheet.merge_range(row, col, row, col + width, b_addr, bold_border_fmt)
        row += 1

        # Empty row for spacing (as per template)
        worksheet.merge_range(row, col, row, col + width, "", bold_border_fmt)
        row += 1

        worksheet.merge_range(
            row, col, row, col + width, f"GSTIN/UIN : {b_gst}", bold_border_fmt
        )
        row += 1

        # Buyer state: only State Name, NO Code (as per template)
        worksheet.merge_range(
            row, col, row, col + width, f"State Name : {b_state}", bold_border_fmt
        )
        row += 1

        worksheet.merge_range(
            row, col, row, col + width, f"Place of Supply : {b_pos}", bold_border_fmt
        )
        row += 1

        return row + 1

    @staticmethod
    def generate_exact_invoice_excel(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate Invoice using 'Invoice_4544.xlsx' as a template.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        import os
        
        template_path = "Invoice_4544.xlsx"
        if not os.path.exists(template_path):
             # Fallback if template missing - log warning and return empty or error
             # For now, assuming it exists as per user input
             logger.error("Template Invoice_4544.xlsx not found.")
             return StreamingResponse(io.BytesIO(b"Template not found"), media_type="text/plain")

        # Load Workbook
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Helper to set cell value safely
        def set_val(coord, value):
            if value is not None:
                ws[coord] = value

        # --- HEADER MAPPING ---
        # Supplier Details (Green - from Settings mostly, but template has Senstographic hardcoded in A3)
        # We will write over it if needed, or assume template is correct for Supplier.
        # User said "Green for data to be fetched from Settings".
        
        # Fetch settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        # If settings exist, overwrite Supplier Block
        if settings.get("supplier_name"):
            set_val("A3", settings["supplier_name"])
        if settings.get("supplier_address"):
            set_val("A4", settings["supplier_address"])
        if settings.get("supplier_gstin"):
            set_val("A5", f"GSTIN/UIN: {settings['supplier_gstin']}")
        if settings.get("supplier_contact"):
            set_val("A7", f"Contact : {settings['supplier_contact']}")
        
        # User Input (Yellow) - Invoice Details
        set_val("L3", header.get("invoice_number", "")) # Invoice No
        set_val("Q3", header.get("invoice_date", ""))   # Date
        set_val("L5", str(header.get("dc_number", "") or "")) # Challan No
        set_val("L6", str(header.get("po_numbers", "") or "")) # PO No
        set_val("Q4", header.get("payment_terms") or "45 Days") # Terms
        set_val("L4", header.get("gem_date") or "") # GEMC Date/Ref if mapped

        # Destination / Dispatch
        set_val("Q7", header.get("srv_date", "")) # SRV Dt
        # Note: SRV No is N7 label, Value likely in Q7 or close. Map says N7 is SRV No label.
        # Template inspect showed [N7] SRV No, [Q7] SRV Dt. Where is SRV No Value? Probably O7 or P7?
        # Let's put SRV No in P7 for now or append to Label?
        # Dump showed: [N7] SRV No. No value in between.
        # We will put SRV No in O7.
        set_val("O7", header.get("srv_number", ""))

        # Buyer Details (Yellow)
        # Fetch Default Buyer if empty
        b_name = header.get("buyer_name")
        b_gst = header.get("buyer_gstin")
        
        if not b_name:
             try:
                buyer = db.execute("SELECT * FROM buyers WHERE is_default=1").fetchone()
                if buyer:
                    b_name = buyer["name"]
                    set_val("A9", b_name)
                    set_val("A10", f"GSTIN/UIN: {buyer['gstin']}")
                    set_val("A11", f"State Name : {buyer.get('state', 'Madhya Pradesh')}")
                    set_val("A12", f"Place of Supply : {buyer.get('place_of_supply', 'BHOPAL, MP')}")
             except: pass
        else:
             set_val("A9", b_name)
             set_val("A10", f"GSTIN/UIN: {b_gst}")
             set_val("A11", f"State Name : {header.get('buyer_state', 'Madhya Pradesh')}")
             set_val("A12", f"Place of Supply : {header.get('place_of_supply', 'BHOPAL, MP')}")

        
        # --- ITEMS (Yellow) ---
        # Start Row 15 (Index 15 means Row 15 in Excel? openpyxl is 1-based, so Row 15)
        start_row = 15
        
        # We need to insert rows if items > 1 (Template has 1 row at 15, then Total at 16)
        # Verify if template already has multiple rows? Map showed A15 data, A16 Total.
        # So we have exactly 1 data row.
        
        num_items = len(items)
        if num_items > 1:
            ws.insert_rows(start_row + 1, amount=num_items - 1)
            # We need to copy styles/merge cells if needed.
            # Ideally, Copy style from Row 15 to new rows.
        
        total_qty = 0
        total_taxable = 0
        total_cgst = 0
        total_sgst = 0
        total_val = 0

        for idx, item in enumerate(items):
            current_row = start_row + idx
            
            # Data Extraction
            qty = float(item.get("quantity", 0) or 0)
            rate = float(item.get("rate", 0) or 0)
            taxable = qty * rate
            
            # Tax Logic (Assuming 9% CGST 9% SGST as per template)
            cgst_rate = 9.0
            sgst_rate = 9.0
            cgst_amt = taxable * (cgst_rate / 100.0)
            sgst_amt = taxable * (sgst_rate / 100.0)
            tot_line = taxable + cgst_amt + sgst_amt

            total_qty += qty
            total_taxable += taxable
            total_cgst += cgst_amt
            total_sgst += sgst_amt
            total_val += tot_line

            # Write Cells
            set_val(f"A{current_row}", item.get("po_item_no", idx + 1)) # PO SL
            set_val(f"B{current_row}", item.get("description", "")) # Desc
            set_val(f"J{current_row}", item.get("material_code", "")) # Mat Code
            set_val(f"L{current_row}", qty)
            set_val(f"M{current_row}", rate)
            set_val(f"N{current_row}", item.get("unit", "NOS"))
            
            # Calculated (Blue)
            set_val(f"O{current_row}", taxable)
            set_val(f"P{current_row}", f"{cgst_rate}%")
            set_val(f"Q{current_row}", cgst_amt)
            set_val(f"R{current_row}", f"{sgst_rate}%")
            set_val(f"S{current_row}", sgst_amt)
            set_val(f"T{current_row}", tot_line)

        # --- TOTALS (Blue) ---
        # The Total Row is now at start_row + num_items
        total_row = start_row + num_items
        # Verify if Total row exists there (it shifts down with insert_rows)
        # Ensure we write to the correct row.
        
        set_val(f"L{total_row}", total_qty)
        set_val(f"O{total_row}", total_taxable)
        set_val(f"Q{total_row}", total_cgst)
        set_val(f"S{total_row}", total_sgst)
        set_val(f"T{total_row}", total_val)
        
        # Words
        amount_words = amount_to_words(total_val)
        set_val(f"A{total_row + 1}", f"Total Amount (In Words):- {amount_words}")

        # Summary Block (Taxable, CGST, SGST, Total) - usually at bottom
        # Map showed [O18] Taxable... [O20]=O16
        # These references need to shift if we added rows.
        # OpenPyXL *should* handle formula shifting if we use insert_rows.
        # But hardcoded values need manual update.
        # We will update the Summary Values explicitly.
        
        # Assuming summary block is fixed relative to Total Row?
        # Map: Total is Row 16. Summary starts Row 18. Gap of 1 row.
        # New Total is `total_row`. Summary starts `total_row + 2`.
        sum_start = total_row + 2
        
        # Update summary values
        # O column: Value
        set_val(f"O{sum_start + 1}", total_taxable) # Taxable Value
        set_val(f"Q{sum_start + 1}", total_cgst)    # CGST Amount
        set_val(f"S{sum_start + 1}", total_sgst)    # SGST Amount
        set_val(f"T{sum_start + 1}", total_cgst + total_sgst) # Total Tax

        # Save to Buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Invoice_{header.get('invoice_number')}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @staticmethod
    def generate_exact_dc_excel(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate strict Excel format matching 'DC12.xls' and User Screenshot
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Delivery Challan")

        # Styles
        workbook.add_format(
            {"bold": True, "font_size": 20, "align": "center", "font_name": "Calibri"}
        )
        border_box = workbook.add_format(
            {
                "border": 1,
                "text_wrap": True,
                "valign": "top",
                "font_name": "Calibri",
                "font_size": 11,
            }
        )
        workbook.add_format(
            {
                "border": 1,
                "bold": True,
                "text_wrap": True,
                "valign": "top",
                "font_name": "Calibri",
                "font_size": 11,
            }
        )
        header_table = workbook.add_format(
            {
                "border": 1,
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "font_name": "Calibri",
                "text_wrap": True,
            }
        )
        cell_fmt = workbook.add_format(
            {"border": 1, "valign": "vcenter", "font_name": "Calibri"}
        )
        cell_center = workbook.add_format(
            {
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "font_name": "Calibri",
            }
        )
        workbook.add_format(
            {
                "border": 1,
                "font_name": "Calibri",
                "font_size": 11,
                "text_wrap": True,
                "valign": "vcenter",
            }
        )  # Added for new layout

        worksheet.set_column("A:A", 10)  # P.O.Sl. No.
        worksheet.set_column("B:B", 60)  # Description
        worksheet.set_column("C:C", 15)  # Quantity

        # Use helper for standardized header
        current_row = ExcelService._write_standard_header(
            worksheet,
            workbook,
            columns=3,
            db=db,
            title="DELIVERY CHALLAN",
            layout="challan",
        )

        # Buyer Block (To...)
        buyer_end_row = ExcelService._write_buyer_block(
            worksheet, workbook, current_row, 0, db, header, width=1
        )

        # Right Header Box (Coincides with Buyer Block)
        worksheet.write(
            current_row, 2, f"Challan No. : {header.get('dc_number', '')}", border_box
        )
        worksheet.write(
            current_row + 1, 2, f"Date : {header.get('dc_date', '')}", border_box
        )
        worksheet.write(
            current_row + 2,
            2,
            f"Your PO No. : {header.get('po_number', '')}",
            border_box,
        )

        # Table Headers
        table_row = max(buyer_end_row, current_row + 4)
        worksheet.write(table_row, 0, "P.O.Sl. No.", header_table)
        worksheet.write(table_row, 1, "Description", header_table)
        worksheet.write(table_row, 2, "Quantity", header_table)

        # Data
        item_row = table_row + 1
        for item in items:
            worksheet.write(item_row, 0, item.get("po_item_no", ""), cell_center)
            worksheet.write(item_row, 1, item.get("description", ""), cell_fmt)
            worksheet.write(
                item_row,
                2,
                f"{item.get('dispatched_quantity', 0)} {item.get('unit', '')}",
                cell_center,
            )
            item_row += 1

        # Fill blank?
        for _ in range(
            item_row, item_row + 5
        ):  # Ensure at least 5 blank rows after items
            worksheet.write(_, 0, "", cell_fmt)
            worksheet.write(_, 1, "", cell_fmt)
            worksheet.write(_, 2, "", cell_fmt)
            item_row += 1

        # Footer
        worksheet.write(item_row, 0, "1", cell_center)

        inv_dt_str = (
            f"Dt. {header.get('invoice_date')}" if header.get("invoice_date") else ""
        )
        worksheet.merge_range(
            item_row,
            1,
            item_row,
            2,
            f"GST Bill No. {header.get('invoice_number', '')} {inv_dt_str}",
            cell_fmt,
        )
        item_row += 1

        worksheet.write(item_row, 0, "2", cell_center)

        gc_dt_str = f"Dt. {header.get('gc_date')}" if header.get("gc_date") else ""
        worksheet.merge_range(
            item_row,
            1,
            item_row,
            2,
            f"Gurantee Certificate No. {header.get('gc_no', '')} {gc_dt_str}",
            cell_fmt,
        )
        item_row += 1

        worksheet.write(item_row, 0, "3", cell_center)
        worksheet.merge_range(
            item_row,
            1,
            item_row,
            2,
            f"Dimension Report {header.get('dr_no', '')}",
            cell_fmt,
        )

        workbook.close()
        output.seek(0)

        filename = f"DC_{header.get('dc_number')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @staticmethod
    def generate_dispatch_summary(
        date_str: str, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate strict Excel format matching 'Summary.xls' and User Screenshot
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Summary")

        # Styles
        workbook.add_format(
            {"bold": True, "font_size": 18, "align": "center", "font_name": "Calibri"}
        )
        workbook.add_format(
            {"bold": True, "font_size": 11, "align": "center", "font_name": "Calibri"}
        )
        header_table = workbook.add_format(
            {
                "bold": True,
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
                "font_name": "Calibri",
            }
        )
        cell_fmt = workbook.add_format(
            {
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "font_name": "Calibri",
            }
        )
        bold_left = workbook.add_format({"bold": True, "font_name": "Calibri"})

        # Column Widths
        worksheet.set_column("A:A", 5)  # S.No.
        worksheet.set_column("B:B", 30)  # Description
        worksheet.set_column("C:C", 15)  # Quantity
        worksheet.set_column("D:D", 8)  # No of packets
        worksheet.set_column("E:E", 12)  # PO NO
        worksheet.set_column("F:F", 18)  # GEMC NO
        worksheet.set_column("G:G", 10)  # Invoice No.
        worksheet.set_column("H:H", 10)  # Challan No.
        worksheet.set_column("I:I", 12)  # Dispatch Delivered

        # Header Section
        current_row = ExcelService._write_standard_header(
            worksheet, workbook, columns=9, db=db, title="SUMMARY", layout="challan"
        )

        worksheet.write(current_row, 0, "Date:", bold_left)
        worksheet.write(current_row, 1, date_str, bold_left)

        table_row = current_row + 2
        headers = [
            "S. No.",
            "Description",
            "Quantity Set/Nos.",
            "No of packets",
            "PO NO",
            "GEMC NO",
            "Invoice No.",
            "Challan No.",
            "Dispatch Delivered",
        ]
        for i, h in enumerate(headers):
            worksheet.write(table_row, i, h, header_table)

        # Data
        row = table_row + 1
        for idx, item in enumerate(items):
            worksheet.write(row, 0, idx + 1, cell_fmt)
            worksheet.write(row, 1, item.get("description", ""), cell_fmt)
            worksheet.write(
                row, 2, f"{item.get('quantity', '')} {item.get('unit', '')}", cell_fmt
            )
            worksheet.write(row, 3, item.get("no_of_packets", ""), cell_fmt)
            worksheet.write(row, 4, item.get("po_number", ""), cell_fmt)
            worksheet.write(row, 5, item.get("gemc_number", ""), cell_fmt)
            worksheet.write(row, 6, item.get("invoice_number", ""), cell_fmt)
            worksheet.write(row, 7, item.get("dc_number", ""), cell_fmt)
            worksheet.write(row, 8, item.get("destination", ""), cell_fmt)
            row += 1

        workbook.close()
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="Summary_{date_str}.xlsx"'
            },
        )

    @staticmethod
    def generate_guarantee_certificate(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate Guarantee Certificate matching the GC5.xls format.
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Guarantee Certificate")

        # Styles
        base_font = "Arial"
        workbook.add_format({"font_name": base_font, "font_size": 12})
        workbook.add_format(
            {"bold": True, "font_size": 16, "align": "center", "font_name": base_font}
        )
        workbook.add_format(
            {"bold": True, "font_size": 10, "align": "center", "font_name": base_font}
        )
        workbook.add_format(
            {"bold": True, "font_size": 14, "align": "center", "font_name": base_font}
        )
        border_all = workbook.add_format(
            {"border": 1, "font_name": base_font, "font_size": 11}
        )
        workbook.add_format(
            {
                "border": 1,
                "text_wrap": True,
                "valign": "top",
                "font_name": base_font,
                "font_size": 11,
            }
        )

        header_table = workbook.add_format(
            {
                "border": 1,
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "font_name": base_font,
                "text_wrap": True,
            }
        )
        cell_fmt = workbook.add_format(
            {"border": 1, "valign": "vcenter", "font_name": base_font}
        )
        cell_center = workbook.add_format(
            {
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "font_name": base_font,
            }
        )
        footer_bold = workbook.add_format(
            {"bold": True, "font_name": base_font, "font_size": 12, "align": "left"}
        )

        # Column Widths
        worksheet.set_column("A:A", 2)
        worksheet.set_column("B:G", 12)
        worksheet.set_column("H:H", 15)
        worksheet.set_column("I:J", 12)

        # Header Section
        current_row = ExcelService._write_standard_header(
            worksheet,
            workbook,
            columns=10,
            db=db,
            title="GUARANTEE  CERTIFICATE",
            layout="challan",
            font_name=base_font,
        )

        # Buyer Block (To...)
        buyer_end_row = ExcelService._write_buyer_block(
            worksheet,
            workbook,
            current_row,
            1,
            db,
            header,
            width=5,
            label="To,",
            font_name=base_font,
        )

        # Info Box (Right Side)
        worksheet.write(current_row, 7, "GC No. & Dt.: ", border_all)

        gc_val = f"{header.get('gc_no', '05')}  dt. {header.get('gc_date', '')}"
        worksheet.merge_range(current_row, 8, current_row, 9, gc_val, border_all)

        worksheet.write(current_row + 1, 7, "PO No. & Dt.: ", border_all)

        po_val = f"{header.get('po_number', '')}  dt. {header.get('po_date', '')}"
        worksheet.merge_range(
            current_row + 1, 8, current_row + 1, 9, po_val, border_all
        )

        worksheet.write(current_row + 2, 7, "DC No. & Dt: ", border_all)

        dc_val = f"{header.get('dc_number', '')}  dt. {header.get('dc_date', '')}"
        worksheet.merge_range(
            current_row + 2, 8, current_row + 2, 9, dc_val, border_all
        )

        # Table Headers
        table_row = max(buyer_end_row, current_row + 4)
        worksheet.write(table_row, 1, "P.O.Sl.\nNo.", header_table)
        worksheet.merge_range(table_row, 2, table_row, 7, "Description", header_table)
        worksheet.merge_range(table_row, 8, table_row, 9, "Quantity", header_table)

        # Data
        item_row = table_row + 1
        for item in items:
            worksheet.write(item_row, 1, item.get("po_item_no", ""), cell_center)
            worksheet.merge_range(
                item_row, 2, item_row, 7, item.get("description", ""), cell_fmt
            )
            worksheet.merge_range(
                item_row,
                8,
                item_row,
                9,
                f"{item.get('quantity', 0)} {item.get('unit', '')}",
                cell_center,
            )
            item_row += 1

        # Blank rows
        while item_row < table_row + 11:
            worksheet.write(item_row, 1, "", cell_center)
            worksheet.merge_range(item_row, 2, item_row, 7, "", cell_fmt)
            worksheet.merge_range(item_row, 8, item_row, 9, "", cell_center)
            item_row += 1

        # Footer Text
        item_row += 1
        footer_text = (
            "The goods supplied as above are guaranteed against manufacturing defects for 24 Month "
            "from delivery date. We undertake to replace or rectify the materials free of cost if any defects "
            "occur during this period."
        )
        worksheet.merge_range(
            item_row,
            1,
            item_row + 2,
            9,
            footer_text,
            workbook.add_format({"text_wrap": True, "font_name": base_font}),
        )
        item_row += 4

        # Signature
        worksheet.write(item_row, 7, "For SENSTOGRAPHIC", footer_bold)

        workbook.close()
        output.seek(0)

        filename = f"GC_{header.get('dc_number')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @staticmethod
    def generate_po_upload_template() -> StreamingResponse:
        """
        Generate empty PO upload template with required headers
        """
        output = io.BytesIO()
        df = pd.DataFrame(
            columns=[
                "PO Number",
                "PO Date",
                "Vendor Name",
                "Project Name",
                "Item No",
                "Material Code",
                "Description",
                "Unit",
                "Qty",
                "Rate",
                "Delivery Date",
            ]
        )
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="PO_Upload")
            workbook = writer.book
            worksheet = writer.sheets["PO_Upload"]
            header_fmt = workbook.add_format(
                {"bold": True, "bg_color": "#D7E4BC", "border": 1}
            )
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_fmt)
            worksheet.set_column("A:K", 15)

        output.seek(0)
        return StreamingResponse(
            output,
            headers={
                "Content-Disposition": 'attachment; filename="PO_Upload_Template.xlsx"'
            },
        )

    @staticmethod
    def generate_srv_upload_template() -> StreamingResponse:
        """
        Generate empty SRV upload template with required headers
        """
        output = io.BytesIO()
        df = pd.DataFrame(
            columns=[
                "SRV Number",
                "SRV Date",
                "PO Number",
                "PO Item No",
                "Lot No",
                "Received Qty",
                "Rejected Qty",
                "Challan No",
                "Challan Date",
                "Invoice No",
                "Invoice Date",
                "Remarks",
            ]
        )
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="SRV_Upload")
            workbook = writer.book
            worksheet = writer.sheets["SRV_Upload"]
            header_fmt = workbook.add_format(
                {"bold": True, "bg_color": "#DDEBF7", "border": 1}
            )
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_fmt)
            worksheet.set_column("A:L", 15)

        output.seek(0)
        return StreamingResponse(
            output,
            headers={
                "Content-Disposition": 'attachment; filename="SRV_Upload_Template.xlsx"'
            },
        )

    @staticmethod
    def generate_exact_dc_excel(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate Delivery Challan using 'Invoice_4544.xlsx' as a template.
        Adapted to show 'DELIVERY CHALLAN'.
        """
        import openpyxl
        import io
        import os
        from fastapi.responses import StreamingResponse

        template_path = "Invoice_4544.xlsx"
        if not os.path.exists(template_path):
             print("Template Invoice_4544.xlsx not found.")
             return StreamingResponse(io.BytesIO(b"Template not found"), media_type="text/plain")

        # Load Workbook
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Helper to set cell value safely
        def set_val(coord, value):
            if value is not None:
                ws[coord] = value

        # --- HEADER TRANSFORMATION ---
        # 1. Change Title
        set_val("A1", "DELIVERY CHALLAN")

        # 2. Supplier Details (from Settings)
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        if settings.get("supplier_name"):
            set_val("A3", settings["supplier_name"])
        if settings.get("supplier_address"):
            set_val("A4", settings["supplier_address"])
        if settings.get("supplier_gstin"):
            set_val("A5", f"GSTIN/UIN: {settings['supplier_gstin']}")
        if settings.get("supplier_contact"):
            set_val("A7", f"Contact : {settings['supplier_contact']}")
        
        # 3. DC Details -> Map to Invoice Slots
        set_val("I3", "Challan No.")
        set_val("L3", header.get("dc_number", "")) 
        
        # Date
        set_val("Q3", header.get("dc_date", ""))

        # Challan No -> PO No?
        set_val("I5", "Order No.")
        set_val("L5", str(header.get("po_numbers", "") or header.get("po_number", "") or ""))

        # Clear other specific Invoice fields
        set_val("I6", "") 
        set_val("L6", "")

        set_val("Q4", header.get("payment_terms") or "") 
        set_val("L4", header.get("gem_date") or "") 

        # Logistics
        set_val("O7", header.get("srv_number", "")) 
        set_val("Q7", header.get("srv_date", ""))

        # Buyer Details
        b_name = header.get("consignee_name")
        if not b_name:
             b_name = header.get("buyer_name")

        set_val("A9", b_name or "")
        set_val("A10", f"GSTIN/UIN: {header.get('consignee_gstin', '')}")
        set_val("A11", f"State: {header.get('destination', '')}") 
        set_val("A12", f"Place of Supply/Dest: {header.get('destination', '')}")

        # --- HEADERS ---
        set_val("L14", "ORDERED") 
        set_val("M14", "DISPATCH")
        set_val("N14", "BALANCE")
        set_val("O14", "RECEIVED")
        set_val("P14", "UNIT")
        set_val("Q14", "HSN")
        # Clear unused headers
        for col in ["R", "S", "T"]:
            set_val(f"{col}14", "")

        # --- ITEMS ---
        start_row = 15
        num_items = len(items)
        if num_items > 1:
            ws.insert_rows(start_row + 1, amount=num_items - 1)
        
        total_qty = 0

        for idx, item in enumerate(items):
            current_row = start_row + idx
            
            # Context quantities
            ord_qty = float(item.get("lot_ordered_qty") or 0)
            disp_qty = float(item.get("dispatched_quantity") or 0)
            bal_qty = float(item.get("remaining_post_dc") or 0)
            recd_qty = float(item.get("received_quantity") or 0)
            
            total_qty += disp_qty

            set_val(f"A{current_row}", item.get("po_item_no", idx + 1))
            set_val(f"B{current_row}", item.get("material_description", "") or item.get("description", ""))
            
            # Map columns per user request: ORDER -> DISPATCH -> BAL -> RECD
            set_val(f"L{current_row}", ord_qty)
            set_val(f"M{current_row}", disp_qty)
            set_val(f"N{current_row}", bal_qty)
            set_val(f"O{current_row}", recd_qty)
            
            set_val(f"P{current_row}", item.get("unit") or "NOS")
            set_val(f"Q{current_row}", item.get("hsn_code", ""))

            # Clear remnants of the Invoice template (Rate/Tax)
            for col in ["R", "S", "T"]:
                set_val(f"{col}{current_row}", "")

        # --- TOTALS ---
        total_row = start_row + num_items
        set_val(f"M{total_row}", total_qty)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"DC_{header.get('dc_number', 'Draft')}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
