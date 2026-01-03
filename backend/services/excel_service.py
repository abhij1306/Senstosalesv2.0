"""
Excel Generation Service
Uses XlsxWriter to generate formatted Excel reports with strict layout control.
"""

import io
import logging
import sqlite3
from typing import Dict, List

import pandas as pd
import xlsxwriter
from fastapi.responses import StreamingResponse

from backend.core.num_to_words import amount_to_words

logger = logging.getLogger(__name__)


class ExcelService:
    @staticmethod
    def generate_response(data: List[Dict], report_type: str) -> StreamingResponse:
        """
        Convert list of dicts to Excel download response (Legacy fallback)
        """
        output = io.BytesIO()

        # Convert to DataFrame
        if not data:
            df = pd.DataFrame()  # Empty
        else:
            df = pd.DataFrame(data)

        # Write to Excel with formatting
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Report", index=False)

            workbook = writer.book
            worksheet = writer.sheets["Report"]

            # Formats
            header_fmt = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#4F81BD",
                    "font_color": "white",
                    "border": 1,
                }
            )

            # Apply header format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_fmt)

            # Auto-adjust column width
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, column_len)

        output.seek(0)

        filename = f"{report_type}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @staticmethod
    def _write_standard_header(
        worksheet,
        workbook,
        columns: int,
        db: sqlite3.Connection,
        title: str = None,
        layout: str = "invoice",
        font_name: str = "Calibri",
    ):
        """
        Consistently writes the business header across all reports with layout options.
        layout: 'invoice' (Standard for Sales Invoice) or 'challan' (Standard for DC, Summary, GC)
        """
        # Fetch settings from DB
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception as e:
            logger.error(f"Failed to fetch business settings, using defaults: {e}")
            settings = {}

        # Default Fallbacks
        # Default Fallbacks - STRICTLY from Settings
        s_name = settings.get("supplier_name", "")
        s_desc = settings.get("supplier_description", "")
        s_addr = settings.get("supplier_address", "")
        s_gst = settings.get("supplier_gstin", "")
        s_phone = settings.get("supplier_contact", "")
        s_state = settings.get("supplier_state", "")
        s_state_code = settings.get("supplier_state_code", "")

        # Formats
        title_fmt = workbook.add_format(
            {"bold": True, "font_size": 18, "align": "center", "font_name": font_name}
        )
        subtitle_fmt = workbook.add_format(
            {"bold": True, "font_size": 10, "align": "center", "font_name": font_name}
        )
        tel_fmt = workbook.add_format({"font_size": 10, "align": "center", "font_name": font_name})
        name_fmt = workbook.add_format(
            {"bold": True, "font_size": 14, "align": "left", "font_name": font_name}
        )
        detail_fmt = workbook.add_format({"font_size": 11, "align": "left", "font_name": font_name})
        bold_detail = workbook.add_format(
            {"bold": True, "font_size": 11, "align": "left", "font_name": font_name}
        )

        row = 0

        if layout == "invoice":
            # Layout matching 'GST_INV_11.xls'
            if title:
                worksheet.merge_range(
                    row,
                    0,
                    row,
                    columns - 1,
                    title,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 14,
                            "align": "center",
                            "font_name": font_name,
                        }
                    ),
                )
                row += 2  # Add spacing

            worksheet.merge_range(row, 0, row, 7, s_name, name_fmt)  # Col H is index 7
            row += 1
            worksheet.merge_range(row, 0, row + 1, 7, s_addr, detail_fmt)
            row += 2
            worksheet.merge_range(row, 0, row, 7, f"GSTIN/UIN: {s_gst}", bold_detail)
            row += 1
            worksheet.merge_range(
                row,
                0,
                row,
                7,
                f"State Name : {s_state}, Code : {s_state_code}",
                bold_detail,
            )
            row += 1
            worksheet.merge_range(row, 0, row, 7, f"Contact : {s_phone}", bold_detail)
            row += 1

        elif layout == "challan":
            # Layout matching 'DC12.xls'
            # Row 1: Tel (Left), GSTIN (Right)
            worksheet.write(row, 0, f"Tel. No. {s_phone}", tel_fmt)
            worksheet.merge_range(row, columns - 1, row, columns - 1, f"GSTIN: {s_gst}", tel_fmt)
            row += 2

            # Branding (Rows 3, 4, 5)
            worksheet.merge_range(row, 0, row, columns - 1, s_name, title_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, columns - 1, s_desc, subtitle_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, columns - 1, s_addr, subtitle_fmt)
            row += 2

            if title:
                worksheet.merge_range(
                    row,
                    0,
                    row,
                    columns - 1,
                    title,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 14,
                            "align": "center",
                            "font_name": font_name,
                        }
                    ),
                )
                row += 1

        return row

    @staticmethod
    def _write_buyer_block(
        worksheet,
        workbook,
        row: int,
        col: int,
        db: sqlite3.Connection,
        header: Dict = None,
        width: int = 5,
        label: str = "Buyer :",
        font_name: str = "Calibri",
    ):
        """
        Consistently writes the Buyer/Consignee block.
        Fetches from DB settings as default, overriden by specific record header if available.
        """
        # Fetch Supplier Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception as e:
            logger.error(f"Failed to fetch business settings for buyer block: {e}")
            settings = {}

        # Fetch Default Buyer if not provided in header
        default_buyer = {}
        if not header.get("consignee_name"):
            try:
                buyer_row = db.execute(
                    "SELECT name, billing_address, gstin, place_of_supply FROM buyers WHERE is_default = 1 AND is_active = 1 LIMIT 1"
                ).fetchone()
                if buyer_row:
                    default_buyer = dict(buyer_row)
            except Exception as e:
                logger.error(f"Failed to fetch default buyer: {e}")

        # Default Buyer Info logic: Header > Default Buyer > Settings > Empty
        b_name = (
            header.get("consignee_name")
            or default_buyer.get("name")
            or settings.get("buyer_name", "")
        )
        b_addr = (
            header.get("consignee_address")
            or default_buyer.get("billing_address")
            or settings.get("buyer_address", "")
        )
        b_gst = (
            header.get("consignee_gstin")
            or default_buyer.get("gstin")
            or settings.get("buyer_gstin", "")
        )

        # Parse Place of Supply if needed
        b_pos_raw = (
            header.get("place_of_supply") or default_buyer.get("place_of_supply") or "BHOPAL, MP"
        )
        b_pos = b_pos_raw

        # State logic (simple extraction if not provided)
        b_state = header.get("buyer_state") or "MP"
        if not header.get("buyer_state") and default_buyer.get("place_of_supply"):
            # Simple heuristic: last word or known states. For now defaulting to MP or extracting from POS
            pass

        # Formats - ALL buyer details should be BOLD with borders
        bold_border_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "font_name": font_name,
                "border": 1,
                "valign": "vcenter",
            }
        )

        if label:
            worksheet.merge_range(row, col, row, col + width, label, bold_border_fmt)
            row += 1

        # Each line should be in its own row with borders - NO merging
        worksheet.merge_range(row, col, row, col + width, b_name, bold_border_fmt)
        row += 1

        # Address might be multi-line but still one row with border
        worksheet.merge_range(row, col, row, col + width, b_addr, bold_border_fmt)
        row += 1

        # Empty row for spacing (as per template)
        worksheet.merge_range(row, col, row, col + width, "", bold_border_fmt)
        row += 1

        worksheet.merge_range(row, col, row, col + width, f"GSTIN/UIN : {b_gst}", bold_border_fmt)
        row += 1

        # Buyer state: only State Name, NO Code (as per template)
        worksheet.merge_range(
            row, col, row, col + width, f"State Name : {b_state}", bold_border_fmt
        )
        row += 1

        worksheet.merge_range(
            row, col, row, col + width, f"Place of Supply : {b_pos}", bold_border_fmt
        )
        row += 1

        return row + 1

    @staticmethod
    def generate_exact_invoice_excel(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate Invoice using 'Invoice_4544.xlsx' as a template.
        Precise Cell Mapping per User Request:
        - Invoice No: L3 | Date: Q3
        - PO No: L6 | Challan No: L5 | Terms: Q4
        - Seller Info: A3-A7
        - Buyer Info: A9-A11 (A12 Place of Supply)
        - Line Items start Row 15: Desc: B15, HSN: I15, Mat Code: J15, Qty: L15, Rate: M15, Unit: N15
        - Totals start Row 16+ (Shifted if items > 1)
        """
        import os
        import openpyxl
        from backend.core.num_to_words import amount_to_words

        template_path = "Invoice_4544.xlsx"
        if not os.path.exists(template_path):
            logger.error("Template Invoice_4544.xlsx not found.")
            return StreamingResponse(io.BytesIO(b"Template not found"), media_type="text/plain")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Helper to set cell value safely and REMOVE reference colors
        from openpyxl.styles import PatternFill
        no_fill = PatternFill(fill_type=None)

        def set_val(coord, value):
            if value is not None:
                ws[coord].value = value
                ws[coord].fill = no_fill

        # Sweep Fill - Remove reference colors from template
        for r_idx in range(1, 40):
            for c_idx in range(1, 26):
                ws.cell(row=r_idx, column=c_idx).fill = no_fill

        # 1. Fetch Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        # 2. Header Info (Yellow)
        set_val("L3", header.get("invoice_number", ""))
        set_val("Q3", header.get("invoice_date", ""))
        set_val("L5", str(header.get("dc_number", "") or ""))
        set_val("L6", str(header.get("po_numbers", "") or ""))
        set_val("Q4", header.get("payment_terms") or "45 Days")
        
        # Optional Logistics (found in screenshot)
        set_val("L4", header.get("gemc_number", ""))
        set_val("Q5", header.get("gemc_date", "")) # Assuming Date next to GEMC
        set_val("P7", header.get("srv_number", ""))
        set_val("Q7", header.get("srv_date", ""))

        # 3. Seller Info (Green) - Explicitly from settings or preserved if missing
        if settings.get("supplier_name"):
            set_val("A3", settings["supplier_name"])
        if settings.get("supplier_address"):
            set_val("A4", settings["supplier_address"])
        if settings.get("supplier_gstin"):
            set_val("A5", f"GSTIN/UIN: {settings['supplier_gstin']}")
        if settings.get("supplier_contact"):
            set_val("A7", f"Contact : {settings['supplier_contact']}")

        # 4. Buyer Info (Yellow/Green)
        b_name = header.get("buyer_name") or header.get("consignee_name")
        b_gst = header.get("buyer_gstin") or header.get("consignee_gstin")
        b_addr = header.get("buyer_address") or header.get("consignee_address")
        
        if b_name:
            set_val("A9", b_name)
            set_val("A10", f"GSTIN/UIN: {b_gst or ''}")
            set_val("A11", f"Address: {b_addr or ''}")
            set_val("A12", f"Place of Supply : {header.get('place_of_supply', 'BHOPAL, MP')}")

        # 5. Line Items (Yellow / Blue Logic)
        start_row = 15
        num_items = len(items)
        if num_items > 1:
            ws.insert_rows(start_row + 1, amount=num_items - 1)
            # Style copying is omitted for brevity as openpyxl insert_rows 
            # might not copy merged cells correctly, but basic data insertion works.

        t_qty = 0
        t_taxable = 0
        t_cgst = 0
        t_sgst = 0
        t_total = 0

        for idx, item in enumerate(items):
            r = start_row + idx
            
            qty = float(item.get("quantity", 0) or 0)
            rate = float(item.get("rate", 0) or 0)
            taxable = qty * rate
            
            # CGST/SGST 9% Logic per Blue requirements
            cgst = taxable * 0.09
            sgst = taxable * 0.09
            line_total = taxable + cgst + sgst
            
            t_qty += qty
            t_taxable += taxable
            t_cgst += cgst
            t_sgst += sgst
            t_total += line_total

            # Data Injection
            set_val(f"A{r}", idx + 1)
            set_val(f"B{r}", item.get("description", ""))
            set_val(f"I{r}", item.get("hsn_sac") or item.get("hsn_code", ""))
            set_val(f"J{r}", item.get("material_code", ""))
            set_val(f"L{r}", qty)
            set_val(f"M{r}", rate)
            set_val(f"N{r}", item.get("unit", "NOS"))
            
            # Blue Logic Calculations
            set_val(f"O{r}", taxable)  # Taxable Value
            set_val(f"P{r}", "9.00%")  # Rate
            set_val(f"Q{r}", cgst)     # Amount
            set_val(f"R{r}", "9.00%")  # Rate
            set_val(f"S{r}", sgst)     # Amount
            set_val(f"T{r}", line_total)

        # 6. Totals & Tax Summary
        # Note: Row indices shift by (num_items - 1)
        shift = num_items - 1
        total_row = 16 + shift
        
        set_val(f"L{total_row}", t_qty)
        set_val(f"O{total_row}", t_taxable)
        set_val(f"Q{total_row}", t_cgst)
        set_val(f"S{total_row}", t_sgst)
        set_val(f"T{total_row}", t_total)

        # Words Conversion
        set_val(f"A{17 + shift}", f"Total Amount (In Words):- {amount_to_words(t_total)}")
        
        # Bottom Tax Summary (Row 20 + shift)
        sum_row = 20 + shift
        set_val(f"O{sum_row}", t_taxable)
        set_val(f"Q{sum_row}", t_cgst)
        set_val(f"S{sum_row}", t_sgst)
        set_val(f"T{sum_row}", t_cgst + t_sgst)

        # Tax in Words (A22, A23 + shift)
        set_val(f"A{22 + shift}", f"CGST (in words) : {amount_to_words(t_cgst)}")
        set_val(f"A{23 + shift}", f"SGST (in words) : {amount_to_words(t_sgst)}")

        # 7. Finalize Stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Invoice_{header.get('invoice_number')}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )


    @staticmethod
    def generate_dispatch_summary(
        date_str: str, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate strict Excel format matching 'Summary.xls' and User Screenshot
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Summary")

        # Styles
        workbook.add_format(
            {"bold": True, "font_size": 18, "align": "center", "font_name": "Calibri"}
        )
        workbook.add_format(
            {"bold": True, "font_size": 11, "align": "center", "font_name": "Calibri"}
        )
        header_table = workbook.add_format(
            {
                "bold": True,
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
                "font_name": "Calibri",
            }
        )
        cell_fmt = workbook.add_format(
            {
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "font_name": "Calibri",
            }
        )
        bold_left = workbook.add_format({"bold": True, "font_name": "Calibri"})

        # Column Widths
        worksheet.set_column("A:A", 5)  # S.No.
        worksheet.set_column("B:B", 30)  # Description
        worksheet.set_column("C:C", 15)  # Quantity
        worksheet.set_column("D:D", 8)  # No of packets
        worksheet.set_column("E:E", 12)  # PO NO
        worksheet.set_column("F:F", 18)  # GEMC NO
        worksheet.set_column("G:G", 10)  # Invoice No.
        worksheet.set_column("H:H", 10)  # Challan No.
        worksheet.set_column("I:I", 12)  # Dispatch Delivered

        # Header Section
        current_row = ExcelService._write_standard_header(
            worksheet, workbook, columns=9, db=db, title="SUMMARY", layout="challan"
        )

        worksheet.write(current_row, 0, "Date:", bold_left)
        worksheet.write(current_row, 1, date_str, bold_left)

        table_row = current_row + 2
        headers = [
            "S. No.",
            "Description",
            "Quantity Set/Nos.",
            "No of packets",
            "PO NO",
            "GEMC NO",
            "Invoice No.",
            "Challan No.",
            "Dispatch Delivered",
        ]
        for i, h in enumerate(headers):
            worksheet.write(table_row, i, h, header_table)

        # Data
        row = table_row + 1
        for idx, item in enumerate(items):
            worksheet.write(row, 0, idx + 1, cell_fmt)
            worksheet.write(row, 1, item.get("description", ""), cell_fmt)
            worksheet.write(row, 2, f"{item.get('quantity', '')} {item.get('unit', '')}", cell_fmt)
            worksheet.write(row, 3, item.get("no_of_packets", ""), cell_fmt)
            worksheet.write(row, 4, item.get("po_number", ""), cell_fmt)
            worksheet.write(row, 5, item.get("gemc_number", ""), cell_fmt)
            worksheet.write(row, 6, item.get("invoice_number", ""), cell_fmt)
            worksheet.write(row, 7, item.get("dc_number", ""), cell_fmt)
            worksheet.write(row, 8, item.get("destination", ""), cell_fmt)
            row += 1

        workbook.close()
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Summary_{date_str}.xlsx"'},
        )

    @staticmethod
    def generate_guarantee_certificate(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate Guarantee Certificate matching the GC5.xls format.
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Guarantee Certificate")

        # Styles
        base_font = "Arial"
        workbook.add_format({"font_name": base_font, "font_size": 12})
        workbook.add_format(
            {"bold": True, "font_size": 16, "align": "center", "font_name": base_font}
        )
        workbook.add_format(
            {"bold": True, "font_size": 10, "align": "center", "font_name": base_font}
        )
        workbook.add_format(
            {"bold": True, "font_size": 14, "align": "center", "font_name": base_font}
        )
        border_all = workbook.add_format({"border": 1, "font_name": base_font, "font_size": 11})
        workbook.add_format(
            {
                "border": 1,
                "text_wrap": True,
                "valign": "top",
                "font_name": base_font,
                "font_size": 11,
            }
        )

        header_table = workbook.add_format(
            {
                "border": 1,
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "font_name": base_font,
                "text_wrap": True,
            }
        )
        cell_fmt = workbook.add_format({"border": 1, "valign": "vcenter", "font_name": base_font})
        cell_center = workbook.add_format(
            {
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "font_name": base_font,
            }
        )
        footer_bold = workbook.add_format(
            {"bold": True, "font_name": base_font, "font_size": 12, "align": "left"}
        )

        # Column Widths
        worksheet.set_column("A:A", 2)
        worksheet.set_column("B:G", 12)
        worksheet.set_column("H:H", 15)
        worksheet.set_column("I:J", 12)

        # Header Section
        current_row = ExcelService._write_standard_header(
            worksheet,
            workbook,
            columns=10,
            db=db,
            title="GUARANTEE  CERTIFICATE",
            layout="challan",
            font_name=base_font,
        )

        # Buyer Block (To...)
        buyer_end_row = ExcelService._write_buyer_block(
            worksheet,
            workbook,
            current_row,
            1,
            db,
            header,
            width=5,
            label="To,",
            font_name=base_font,
        )

        # Info Box (Right Side)
        worksheet.write(current_row, 7, "GC No. & Dt.: ", border_all)

        gc_val = f"{header.get('gc_no', '05')}  dt. {header.get('gc_date', '')}"
        worksheet.merge_range(current_row, 8, current_row, 9, gc_val, border_all)

        worksheet.write(current_row + 1, 7, "PO No. & Dt.: ", border_all)

        po_val = f"{header.get('po_number', '')}  dt. {header.get('po_date', '')}"
        worksheet.merge_range(current_row + 1, 8, current_row + 1, 9, po_val, border_all)

        worksheet.write(current_row + 2, 7, "DC No. & Dt: ", border_all)

        dc_val = f"{header.get('dc_number', '')}  dt. {header.get('dc_date', '')}"
        worksheet.merge_range(current_row + 2, 8, current_row + 2, 9, dc_val, border_all)

        # Table Headers
        table_row = max(buyer_end_row, current_row + 4)
        worksheet.write(table_row, 1, "P.O.Sl.\nNo.", header_table)
        worksheet.merge_range(table_row, 2, table_row, 7, "Description", header_table)
        worksheet.merge_range(table_row, 8, table_row, 9, "Quantity", header_table)

        # Data
        item_row = table_row + 1
        for item in items:
            worksheet.write(item_row, 1, item.get("po_item_no", ""), cell_center)
            worksheet.merge_range(item_row, 2, item_row, 7, item.get("description", ""), cell_fmt)
            worksheet.merge_range(
                item_row,
                8,
                item_row,
                9,
                f"{item.get('quantity', 0)} {item.get('unit', '')}",
                cell_center,
            )
            item_row += 1

        # Blank rows
        while item_row < table_row + 11:
            worksheet.write(item_row, 1, "", cell_center)
            worksheet.merge_range(item_row, 2, item_row, 7, "", cell_fmt)
            worksheet.merge_range(item_row, 8, item_row, 9, "", cell_center)
            item_row += 1

        # Footer Text
        item_row += 1
        footer_text = (
            "The goods supplied as above are guaranteed against manufacturing defects for 24 Month "
            "from delivery date. We undertake to replace or rectify the materials free of cost if any defects "
            "occur during this period."
        )
        worksheet.merge_range(
            item_row,
            1,
            item_row + 2,
            9,
            footer_text,
            workbook.add_format({"text_wrap": True, "font_name": base_font}),
        )
        item_row += 4

        # Signature
        worksheet.write(item_row, 7, "For SENSTOGRAPHIC", footer_bold)

        workbook.close()
        output.seek(0)

        filename = f"GC_{header.get('dc_number')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @staticmethod
    def generate_po_upload_template() -> StreamingResponse:
        """
        Generate empty PO upload template with required headers
        """
        output = io.BytesIO()
        df = pd.DataFrame(
            columns=[
                "PO Number",
                "PO Date",
                "Vendor Name",
                "Project Name",
                "Item No",
                "Material Code",
                "Description",
                "Unit",
                "Qty",
                "Rate",
                "Delivery Date",
            ]
        )
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="PO_Upload")
            workbook = writer.book
            worksheet = writer.sheets["PO_Upload"]
            header_fmt = workbook.add_format({"bold": True, "bg_color": "#D7E4BC", "border": 1})
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_fmt)
            worksheet.set_column("A:K", 15)

        output.seek(0)
        return StreamingResponse(
            output,
            headers={"Content-Disposition": 'attachment; filename="PO_Upload_Template.xlsx"'},
        )

    @staticmethod
    def generate_srv_upload_template() -> StreamingResponse:
        """
        Generate empty SRV upload template with required headers
        """
        output = io.BytesIO()
        df = pd.DataFrame(
            columns=[
                "SRV Number",
                "SRV Date",
                "PO Number",
                "PO Item No",
                "Lot No",
                "Received Qty",
                "Rejected Qty",
                "Challan No",
                "Challan Date",
                "Invoice No",
                "Invoice Date",
                "Remarks",
            ]
        )
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="SRV_Upload")
            workbook = writer.book
            worksheet = writer.sheets["SRV_Upload"]
            header_fmt = workbook.add_format({"bold": True, "bg_color": "#DDEBF7", "border": 1})
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_fmt)
            worksheet.set_column("A:L", 15)

        output.seek(0)
        return StreamingResponse(
            output,
            headers={"Content-Disposition": 'attachment; filename="SRV_Upload_Template.xlsx"'},
        )

    @staticmethod
    def generate_exact_dc_excel(
        header: Dict, items: List[Dict], db: sqlite3.Connection
    ) -> StreamingResponse:
        """
        Generate Delivery Challan using 'Invoice_4544.xlsx' as a template.
        Adapted to show 'DELIVERY CHALLAN'.
        """
        import io
        import os

        import openpyxl
        from fastapi.responses import StreamingResponse

        template_path = "Invoice_4544.xlsx"
        if not os.path.exists(template_path):
            print("Template Invoice_4544.xlsx not found.")
            return StreamingResponse(io.BytesIO(b"Template not found"), media_type="text/plain")

        # Load Workbook
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Helper to set cell value safely and REMOVE reference colors
        from openpyxl.styles import PatternFill
        no_fill = PatternFill(fill_type=None)

        def set_val(coord, value):
            if value is not None:
                ws[coord].value = value
                ws[coord].fill = no_fill

        # Sweep Fill - Remove reference colors from template
        for r_idx in range(1, 40):
            for c_idx in range(1, 26):
                ws.cell(row=r_idx, column=c_idx).fill = no_fill

        # --- HEADER TRANSFORMATION ---
        # 1. Change Title
        set_val("A1", "DELIVERY CHALLAN")

        # 2. Supplier Details (from Settings)
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        if settings.get("supplier_name"):
            set_val("A3", settings["supplier_name"])
        if settings.get("supplier_address"):
            set_val("A4", settings["supplier_address"])
        if settings.get("supplier_gstin"):
            set_val("A5", f"GSTIN/UIN: {settings['supplier_gstin']}")
        if settings.get("supplier_contact"):
            set_val("A7", f"Contact : {settings['supplier_contact']}")

        # 3. DC Details -> Map to Invoice Slots
        set_val("I3", "Challan No.")
        set_val("L3", header.get("dc_number", ""))

        # Date
        set_val("Q3", header.get("dc_date", ""))

        # Challan No -> PO No?
        set_val("I5", "Order No.")
        set_val("L5", str(header.get("po_numbers", "") or header.get("po_number", "") or ""))

        # Clear other specific Invoice fields
        set_val("I6", "")
        set_val("L6", "")

        set_val("Q4", header.get("payment_terms") or "")
        set_val("L4", header.get("gem_date") or "")

        # Logistics
        set_val("O7", header.get("srv_number", ""))
        set_val("Q7", header.get("srv_date", ""))

        # Buyer Details
        b_name = header.get("consignee_name")
        if not b_name:
            b_name = header.get("buyer_name")

        set_val("A9", b_name or "")
        set_val("A10", f"GSTIN/UIN: {header.get('consignee_gstin', '')}")
        set_val("A11", f"State: {header.get('destination', '')}")
        set_val("A12", f"Place of Supply/Dest: {header.get('destination', '')}")

        # --- HEADERS ---
        set_val("L14", "ORDERED")
        set_val("M14", "DISPATCH")
        set_val("N14", "BALANCE")
        set_val("O14", "RECEIVED")
        set_val("P14", "UNIT")
        set_val("Q14", "HSN")
        # Clear unused headers
        for col in ["R", "S", "T"]:
            set_val(f"{col}14", "")

        # --- ITEMS ---
        start_row = 15
        num_items = len(items)
        if num_items > 1:
            ws.insert_rows(start_row + 1, amount=num_items - 1)

        total_qty = 0

        for idx, item in enumerate(items):
            current_row = start_row + idx

            # Context quantities
            ord_qty = float(item.get("lot_ordered_qty") or 0)
            disp_qty = float(item.get("dispatched_quantity") or 0)
            bal_qty = float(item.get("remaining_post_dc") or 0)
            recd_qty = float(item.get("received_quantity") or 0)

            total_qty += disp_qty

            set_val(f"A{current_row}", item.get("po_item_no", idx + 1))
            set_val(
                f"B{current_row}",
                item.get("material_description", "") or item.get("description", ""),
            )

            # Map columns per user request: ORDER -> DISPATCH -> BAL -> RECD
            set_val(f"L{current_row}", ord_qty)
            set_val(f"M{current_row}", disp_qty)
            set_val(f"N{current_row}", bal_qty)
            set_val(f"O{current_row}", recd_qty)

            set_val(f"P{current_row}", item.get("unit") or "NOS")
            set_val(f"Q{current_row}", item.get("hsn_code", ""))

            # Clear remnants of the Invoice template (Rate/Tax)
            for col in ["R", "S", "T"]:
                set_val(f"{col}{current_row}", "")

        # --- TOTALS ---
        total_row = start_row + num_items
        set_val(f"M{total_row}", total_qty)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"DC_{header.get('dc_number', 'Draft')}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
